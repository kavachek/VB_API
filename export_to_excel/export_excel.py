import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference


def add_chart_and_data(ws, data_start_col, categories_start_col):
    """Добавляет данные и график в рабочий лист."""
    chart = BarChart()
    chart.title = "Склад"
    chart.y_axis.title = "Количество заказов"

    data = Reference(ws, min_col=data_start_col, min_row=1, max_row=ws.max_row, max_col=data_start_col)
    categories = Reference(ws, min_col=categories_start_col, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "H2")


def export_sales_report(sales_data, save_path, start_date, end_date):
    if not sales_data: return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"

    headers = ["Склад", "Количество заказов", "Общая выручка", "Средняя цена"]
    ws.append(headers)

    orders_by_city = sales_data.get('orders_by_city')
    prices_by_city = sales_data.get('prices_by_city')

    if orders_by_city is None or prices_by_city is None: return None

    merged_sales = orders_by_city.merge(prices_by_city, on='warehouseName', how='left')

    for i, row in merged_sales.iterrows(): ws.append([row['warehouseName'], row['order_count'],
                                                      row['total_revenue'], row['average_price']])

    # Добавляем данные и график с использованием общей функции
    add_chart_and_data(ws, 2, 1)

    filename = f"Продажи {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}.xlsx"
    full_path = os.path.join(save_path, filename)
    wb.save(full_path)


def export_stocks_report(stocks_data, save_path, start_date, end_date):
    if not stocks_data: return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"

    headers = ["Склад", "Общий остаток"]
    ws.append(headers)

    stocks_by_city = stocks_data.get('stocks_by_city')

    if stocks_by_city is None: return None

    for i, row in stocks_by_city.iterrows(): ws.append([row['warehouseName'], row['total_stock']])

    # Добавляем данные и график с использованием общей функции
    add_chart_and_data(ws, 2, 1)

    filename = f"Остатки {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}.xlsx"
    full_path = os.path.join(save_path, filename)
    wb.save(full_path)
